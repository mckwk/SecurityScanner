import time
import psutil
import os
from openpyxl import Workbook, load_workbook
from ttkthemes import ThemedTk
from UI.gui import GUI
from datetime import datetime
import config

def measure_time(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"{func.__name__} execution time: {end_time - start_time:.2f} seconds")
        return result, end_time - start_time
    return wrapper

@measure_time
def initialize_gui():
    root = ThemedTk(theme="arc")
    app = GUI(root)
    return root, app

def start_main_loop(root):
    root.mainloop()

def measure_memory_usage():
    process = psutil.Process(os.getpid())
    memory_info = process.memory_info()
    return memory_info.rss

def measure_cpu_usage(duration=5, interval=1):
    process = psutil.Process(os.getpid())
    max_cpu_percent = 0
    for _ in range(int(duration / interval)):
        cpu_percent = process.cpu_percent(interval=interval) / psutil.cpu_count()
        if cpu_percent > max_cpu_percent:
            max_cpu_percent = cpu_percent
    return max_cpu_percent

def measure_network_usage():
    net_io = psutil.net_io_counters(pernic=False)
    return net_io.bytes_sent, net_io.bytes_recv

def measure_network_scan(app):
    start_time = time.time()
    app.network_scanner.scan_network()
    end_time = time.time()
    network_scan_time = end_time - start_time
    print(f"Network scan execution time: {network_scan_time:.2f} seconds")
    return network_scan_time

def measure_vulnerability_search(app):
    start_time = time.time()
    app.vulnerability_checker.search_vulnerabilities("test_model", "test_vendor")
    end_time = time.time()
    vulnerability_search_time = end_time - start_time
    print(f"Vulnerability search execution time: {vulnerability_search_time:.2f} seconds")
    return vulnerability_search_time

def append_to_excel(file_path, data):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        next_run = sheet.max_row
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Timestamp", "Run no.", "Init Time (s)", "Network Scan Time (s)", "Vuln Search Time (s)", "Memory Usage (MB)", "CPU Percent (%)", "Data Sent (KB)", "Data Received (KB)"])
        next_run = 1 

    data[0] = next_run 
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp] + data)
    workbook.save(file_path)

if __name__ == "__main__":
    import sys

    if len(sys.argv) != 2:
        print("Usage: python performance_metrics.py <number_of_runs>")
        sys.exit(1)

    num_runs = int(sys.argv[1])

    for run in range(1, num_runs + 1):
        print(f"Starting run {run}/{num_runs}")

        # Measure initialization time
        (root, app), init_time = initialize_gui()

        # Capture initial network usage
        initial_bytes_sent, initial_bytes_recv = measure_network_usage()

        # Measure specific aspects of the program
        network_scan_time = measure_network_scan(app) # seconds
        vulnerability_search_time = measure_vulnerability_search(app) # seconds
        mem_usage = measure_memory_usage() / (1024 * 1024) # MB
        cpu_percent = measure_cpu_usage() # %

        # Capture final network usage
        final_bytes_sent, final_bytes_recv = measure_network_usage()    
        bytes_sent = (final_bytes_sent - initial_bytes_sent) / 1024 # KB
        bytes_received = (final_bytes_recv - initial_bytes_recv) / 1024 # KB

        data = [None, init_time, network_scan_time, vulnerability_search_time, mem_usage, cpu_percent, bytes_sent, bytes_received]
        append_to_excel(config.EXCEL_PERF_FILE, data)

        root.destroy()
        print(f"Run {run}/{num_runs} completed")
